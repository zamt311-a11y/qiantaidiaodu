from __future__ import annotations

import csv
import io
import json
import os
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_admin
from app.core.config import settings
from app.db.session import get_db
from app.models.task import Task
from app.models.task_photo import TaskPhoto
from app.models.user import User
from app.schemas.task import TaskAdminUpdate, TaskCreate, TaskOut, TaskUpdate
from app.utils.geo import haversine_distance_m


router = APIRouter()


class TaskDispatchRequest(BaseModel):
    task_ids: list[int]
    assignee_id: int


class TaskBulkDeleteRequest(BaseModel):
    task_ids: list[int]


class ImportErrorItem(BaseModel):
    row: int
    error: str


class ImportPreview(BaseModel):
    headers: list[str]
    sample_rows: list[dict[str, str]]


def _norm_header(s: str) -> str:
    return "".join(ch for ch in s.strip().lower().replace(" ", "").replace("_", "") if ch not in "\ufeff")


def _pick(row: dict[str, str], keys: list[str]) -> str | None:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return str(row[k]).strip()
    return None


def _mapping_keys(mapping_json: str | None, fallback: list[str]) -> list[str]:
    if not mapping_json:
        return [_norm_header(k) for k in fallback]
    try:
        m = json.loads(mapping_json)
    except json.JSONDecodeError:
        return [_norm_header(k) for k in fallback]
    if not isinstance(m, dict):
        return [_norm_header(k) for k in fallback]
    return [_norm_header(str(k)) for k in m if str(k).strip()]


def _parse_dt(value: str | None) -> datetime | None:
    if value is None or value == "":
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _read_csv(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("gb18030")
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows = list(reader)
    if not rows:
        return ([], [])
    headers = [_norm_header(h) for h in rows[0]]
    data: list[dict[str, str]] = []
    for r in rows[1:]:
        if not any(cell.strip() for cell in r if isinstance(cell, str)):
            continue
        item: dict[str, str] = {}
        for idx, h in enumerate(headers):
            item[h] = r[idx].strip() if idx < len(r) else ""
        data.append(item)
    return (headers, data)


def _read_excel(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ([], [])
    headers = [_norm_header(str(c or "")) for c in rows[0]]
    data: list[dict[str, str]] = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        item: dict[str, str] = {}
        for idx, h in enumerate(headers):
            v = r[idx] if idx < len(r) else ""
            item[h] = "" if v is None else str(v).strip()
        data.append(item)
    return (headers, data)


def _preview_csv(file_bytes: bytes, sample_size: int = 5) -> ImportPreview:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("gb18030")
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows = list(reader)
    if not rows:
        return ImportPreview(headers=[], sample_rows=[])
    headers_raw = [str(h or "").strip() for h in rows[0]]
    sample_rows_out: list[dict[str, str]] = []
    for r in rows[1:]:
        if len(sample_rows_out) >= sample_size:
            break
        if not any(str(cell or "").strip() for cell in r):
            continue
        item: dict[str, str] = {}
        for idx, h in enumerate(headers_raw):
            item[h] = str(r[idx]).strip() if idx < len(r) and r[idx] is not None else ""
        sample_rows_out.append(item)
    return ImportPreview(headers=headers_raw, sample_rows=sample_rows_out)


def _preview_excel(file_bytes: bytes, sample_size: int = 5) -> ImportPreview:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ImportPreview(headers=[], sample_rows=[])
    headers_raw = [str(c).strip() if c is not None else "" for c in rows[0]]
    sample_rows_out: list[dict[str, str]] = []
    for r in rows[1:]:
        if len(sample_rows_out) >= sample_size:
            break
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        item: dict[str, str] = {}
        for idx, h in enumerate(headers_raw):
            v = r[idx] if idx < len(r) else ""
            item[h] = "" if v is None else str(v).strip()
        sample_rows_out.append(item)
    return ImportPreview(headers=headers_raw, sample_rows=sample_rows_out)


@router.get("", response_model=list[TaskOut])
def list_tasks(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    status_: str | None = Query(default=None, alias="status"),
    task_type: str | None = None,
    priority: str | None = None,
    assignee_id: int | None = None,
    assignee_none: bool | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    near_lat: float | None = None,
    near_lon: float | None = None,
    radius_km: float | None = None,
    q: str | None = None,
    offset: int = 0,
    limit: int | None = None,
) -> list[TaskOut]:
    stmt = select(Task)
    if current_user.role != "admin":
        stmt = stmt.where(Task.assignee_id == current_user.id)
    if status_:
        stmt = stmt.where(Task.status == status_)
    if task_type:
        stmt = stmt.where(Task.task_type == task_type)
    if priority:
        stmt = stmt.where(Task.priority == priority)
    if assignee_id is not None:
        stmt = stmt.where(Task.assignee_id == assignee_id)
    if assignee_none is True:
        stmt = stmt.where(Task.assignee_id.is_(None))
    df = _parse_dt(date_from)
    dt = _parse_dt(date_to)
    if df:
        stmt = stmt.where(Task.planned_start_at >= df)
    if dt:
        stmt = stmt.where(Task.planned_start_at <= dt)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Task.site_id.ilike(like),
                Task.site_name.ilike(like),
                Task.address.ilike(like),
                Task.remark.ilike(like),
            )
        )
    stmt = stmt.order_by(Task.id.desc())
    if offset < 0:
        offset = 0
    if limit is not None:
        if limit < 1:
            limit = 1
        if limit > 5000:
            limit = 5000
        stmt = stmt.offset(offset).limit(limit)
    tasks = list(db.scalars(stmt).all())
    if near_lat is not None and near_lon is not None and radius_km is not None:
        radius_m = radius_km * 1000.0
        tasks = [
            t
            for t in tasks
            if haversine_distance_m(near_lat, near_lon, t.lat, t.lon) <= radius_m
        ]
    return [TaskOut.model_validate(t) for t in tasks]


@router.post("", response_model=TaskOut)
def create_task(
    payload: TaskCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> TaskOut:
    task = Task(
        site_id=payload.site_id,
        site_name=payload.site_name or "",
        lon=float(payload.lon),
        lat=float(payload.lat),
        task_type=payload.task_type or "",
        priority=payload.priority or "中",
        status=payload.status or "待执行",
        planned_start_at=_parse_dt(payload.planned_start_at),
        planned_end_at=_parse_dt(payload.planned_end_at),
        address=payload.address or "",
        remark=payload.remark or "",
        assignee_id=payload.assignee_id,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return TaskOut.model_validate(task)


@router.get("/{task_id}", response_model=TaskOut)
def get_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> TaskOut:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="任务不存在")
    if current_user.role != "admin" and task.assignee_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限")
    return TaskOut.model_validate(task)


@router.patch("/{task_id}", response_model=TaskOut)
def update_task(
    task_id: int,
    payload: TaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> TaskOut:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="任务不存在")
    if current_user.role != "admin" and task.assignee_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限")
    if payload.status is not None:
        task.status = payload.status
    if payload.remark is not None:
        task.remark = payload.remark
    db.add(task)
    db.commit()
    db.refresh(task)
    return TaskOut.model_validate(task)


@router.patch("/{task_id}/admin", response_model=TaskOut)
def admin_update_task(
    task_id: int,
    payload: TaskAdminUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> TaskOut:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="任务不存在")

    if payload.site_id is not None:
        task.site_id = payload.site_id
    if payload.site_name is not None:
        task.site_name = payload.site_name
    if payload.lon is not None:
        task.lon = float(payload.lon)
    if payload.lat is not None:
        task.lat = float(payload.lat)
    if payload.task_type is not None:
        task.task_type = payload.task_type
    if payload.priority is not None:
        task.priority = payload.priority
    if payload.status is not None:
        task.status = payload.status
    if payload.planned_start_at is not None:
        task.planned_start_at = _parse_dt(payload.planned_start_at)
    if payload.planned_end_at is not None:
        task.planned_end_at = _parse_dt(payload.planned_end_at)
    if payload.address is not None:
        task.address = payload.address
    if payload.remark is not None:
        task.remark = payload.remark
    if "assignee_id" in payload.model_fields_set:
        task.assignee_id = payload.assignee_id

    db.add(task)
    db.commit()
    db.refresh(task)
    return TaskOut.model_validate(task)


@router.delete("/{task_id}")
def delete_task(
    task_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="任务不存在")
    db.delete(task)
    db.commit()
    return {"deleted": 1}


@router.post("/bulk_delete")
def bulk_delete_tasks(
    payload: TaskBulkDeleteRequest,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    if not payload.task_ids:
        return {"deleted": 0}
    tasks = list(db.scalars(select(Task).where(Task.id.in_(payload.task_ids))).all())
    for t in tasks:
        db.delete(t)
    db.commit()
    return {"deleted": len(tasks)}


@router.post("/dispatch")
def dispatch_tasks(
    payload: TaskDispatchRequest,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    assignee = db.get(User, payload.assignee_id)
    if assignee is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="负责人不存在")
    tasks = list(db.scalars(select(Task).where(Task.id.in_(payload.task_ids))).all())
    for t in tasks:
        t.assignee_id = assignee.id
        if t.status == "":
            t.status = "待执行"
        db.add(t)
    db.commit()
    return {"updated": len(tasks)}


@router.post("/import")
def import_tasks(
    file: UploadFile = File(...),
    mapping_json: str | None = Form(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    content = file.file.read()
    filename = file.filename or ""
    if filename.lower().endswith(".xlsx"):
        headers, rows = _read_excel(content)
    else:
        headers, rows = _read_csv(content)

    default_site_id = ["站点id", "siteid", "site_id", "站点编号"]
    default_site_name = ["站点名称", "sitename", "name"]
    default_lon = ["经度", "站点经度", "lon", "longitude", "lng"]
    default_lat = ["纬度", "站点纬度", "lat", "latitude"]
    default_task_type = ["任务类型", "tasktype", "type"]
    default_priority = ["优先级", "priority"]
    default_status = ["状态", "status"]
    default_start = ["计划开始时间", "计划开始", "starttime", "plannedstart"]
    default_end = ["计划完成时间", "计划完成", "endtime", "plannedend"]
    default_address = ["地址", "address"]
    default_remark = ["备注", "remark"]

    mapping: dict[str, str] = {}
    if mapping_json:
        try:
            mapping = json.loads(mapping_json)
        except json.JSONDecodeError:
            mapping = {}
    if not isinstance(mapping, dict):
        mapping = {}

    def k(field: str, fallback: list[str]) -> list[str]:
        v = mapping.get(field)
        if isinstance(v, str) and v.strip():
            return [_norm_header(v)]
        return [_norm_header(x) for x in fallback]

    errors: list[ImportErrorItem] = []
    inserted = 0
    for idx, r in enumerate(rows, start=2):
        site_id = _pick(r, k("site_id", default_site_id))
        lon_s = _pick(r, k("lon", default_lon))
        lat_s = _pick(r, k("lat", default_lat))
        if not site_id:
            errors.append(ImportErrorItem(row=idx, error="缺少站点ID"))
            continue
        try:
            lon = float(lon_s) if lon_s is not None else None
            lat = float(lat_s) if lat_s is not None else None
        except ValueError:
            lon = None
            lat = None
        if lon is None or lat is None:
            errors.append(ImportErrorItem(row=idx, error="经纬度格式错误"))
            continue
        task = Task(
            site_id=site_id,
            site_name=_pick(r, k("site_name", default_site_name)) or "",
            lon=lon,
            lat=lat,
            task_type=_pick(r, k("task_type", default_task_type)) or "",
            priority=_pick(r, k("priority", default_priority)) or "中",
            status=_pick(r, k("status", default_status)) or "待执行",
            planned_start_at=_parse_dt(_pick(r, k("planned_start_at", default_start))),
            planned_end_at=_parse_dt(_pick(r, k("planned_end_at", default_end))),
            address=_pick(r, k("address", default_address)) or "",
            remark=_pick(r, k("remark", default_remark)) or "",
        )
        db.add(task)
        inserted += 1
    db.commit()
    return {"inserted": inserted, "errors": [e.model_dump() for e in errors]}


@router.post("/import/preview", response_model=ImportPreview)
def preview_task_import(
    file: UploadFile = File(...),
    _: User = Depends(require_admin),
) -> ImportPreview:
    content = file.file.read()
    filename = file.filename or ""
    if filename.lower().endswith(".xlsx"):
        return _preview_excel(content)
    return _preview_csv(content)


@router.post("/{task_id}/photos")
def upload_task_photos(
    task_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(Task, task_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="任务不存在")
    if current_user.role != "admin" and task.assignee_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限")
    if len(files) > 3:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="最多上传3张照片")

    root = settings.resolved_task_photos_dir
    root.mkdir(parents=True, exist_ok=True)
    saved: list[str] = []
    for f in files:
        ext = os.path.splitext(f.filename or "")[1].lower()
        if ext not in (".jpg", ".jpeg", ".png", ".webp"):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持jpg/png/webp")
        name = f"task{task_id}_u{current_user.id}_{int(datetime.utcnow().timestamp()*1000)}{ext}"
        abs_path = (root / name).resolve()
        data = f.file.read()
        abs_path.write_bytes(data)
        rel_path = str(Path(settings.file_dir_task_photos) / name).replace("\\", "/")
        db.add(TaskPhoto(task_id=task_id, uploader_id=current_user.id, file_path=rel_path))
        saved.append(rel_path)
    db.commit()
    return {"saved": saved}

